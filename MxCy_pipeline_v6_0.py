#!/usr/bin/env python3
# =============================================================================
#  MxCy_pipeline_v6_0.py
#  Standalone, cron/scheduler-ready script version of MxCy_pipeline_v6_0.ipynb
#
#  Physics-constrained ML pipeline for transition-metal chalcogenides (MxCy).
#  27-feature physics descriptor set, Materials Project + literature PDF
#  mining, domain-split gap ensembles, ordering-aware magnetisation,
#  bootstrap-calibrated 90% confidence intervals.
#
#  Designed to run unattended:  python MxCy_pipeline_v6_0.py
#  Logs progress to stdout AND to a timestamped .log file in OUTPUT_DIR.
#
#  Author: Abderrahmane REGGAD
# =============================================================================

from __future__ import annotations

import sys
import os
import re
import math
import json
import time
import logging
import warnings
import subprocess
from pathlib import Path
from datetime import datetime
from itertools import product
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import pandas as pd
import joblib

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from sklearn.ensemble import (
    RandomForestRegressor, RandomForestClassifier,
    GradientBoostingRegressor, HistGradientBoostingRegressor,
    ExtraTreesRegressor,
)
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import (
    KFold, StratifiedKFold, StratifiedShuffleSplit, RandomizedSearchCV,
)
from sklearn.calibration import CalibratedClassifierCV
from sklearn.isotonic import IsotonicRegression
from sklearn.metrics import r2_score, mean_absolute_error, accuracy_score
from sklearn.multioutput import MultiOutputRegressor

import xgboost as xgb

try:
    import shap
    SHAP_OK = True
except ImportError:
    SHAP_OK = False

warnings.filterwarnings("ignore")

# =============================================================================
#  CONFIGURATION — edit these before running
# =============================================================================

# ── Materials Project API key ───────────────────────────────────────────────
# Best practice for a scheduled/cron script: set as an environment variable
#   export MP_API_KEY="your_key_here"
# rather than hardcoding it here.
MP_API_KEY = os.environ.get("MP_API_KEY", "")

# ── Output directory ─────────────────────────────────────────────────────────
OUTPUT_DIR = os.environ.get("MXCY_OUTPUT_DIR", "./MxCy_outputs_v6")

# ── Literature PDF directory (Google Drive, local folder, etc.) ─────────────
# Examples:
#   Windows local:     r"G:\MyResearch\AllPapers"
#   Linux/macOS local: "/home/user/papers"
#   Google Drive (rclone-mounted or synced locally):
#                       "/home/user/GoogleDrive/MyResearch/AllPapers"
#   Google Drive (via Colab-style mount, if running interactively):
#                       "/content/drive/MyDrive/AllPapers"
PAPER_DIR = os.environ.get("MXCY_PAPER_DIR", "")

# ── Optional manually curated CSV of paper-extracted values ─────────────────
# Columns required: compound, M, C, structure, band_gap_eV, magnetization_muB
PAPER_DB_CSV_PATH = os.environ.get("MXCY_PAPER_CSV", "")

# ── PDF mining performance tuning (for 1000+ paper libraries) ───────────────
PDF_MAX_WORKERS = int(os.environ.get("MXCY_PDF_WORKERS", "8"))   # parallel threads
PDF_BATCH_SIZE = int(os.environ.get("MXCY_PDF_BATCH", "50"))     # progress report interval
PDF_TIMEOUT_SEC = int(os.environ.get("MXCY_PDF_TIMEOUT", "45"))  # per-file timeout
PDF_SKIP_LARGER_THAN_MB = float(os.environ.get("MXCY_PDF_MAX_MB", "60"))  # skip huge scans

# =============================================================================
#  LOGGING SETUP — stdout + timestamped log file (cron-friendly)
# =============================================================================

os.makedirs(OUTPUT_DIR, exist_ok=True)
_log_path = os.path.join(
    OUTPUT_DIR, f"mxcy_run_{datetime.now():%Y%m%d_%H%M%S}.log"
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(_log_path, encoding="utf-8"),
    ],
)
log = logging.getLogger("mxcy")
log.info("=" * 70)
log.info("MxCy Pipeline v6.0 — standalone script run starting")
log.info("Log file: %s", _log_path)
log.info("=" * 70)

# =============================================================================
#  ATOMIC & PHYSICS CONSTANTS  (v6.0: 27-feature set)
# =============================================================================

ATOMIC_DATA = {
    # 3d:  [n_d, r_cov, chi, IE, U_Hubbard, r_ionic]
    "Ti": [2,  1.40, 1.54, 6.82, 3.0, 0.86],
    "V":  [3,  1.35, 1.63, 6.74, 3.5, 0.79],
    "Cr": [4,  1.28, 1.66, 6.77, 3.5, 0.87],
    "Mn": [5,  1.27, 1.55, 7.43, 4.0, 0.97],
    "Fe": [6,  1.26, 1.83, 7.90, 4.5, 0.92],
    "Co": [7,  1.25, 1.88, 7.88, 5.0, 0.88],
    "Ni": [8,  1.24, 1.91, 7.64, 5.5, 0.83],
    "Cu": [9,  1.28, 1.90, 7.73, 7.0, 0.87],
    "Zn": [10, 1.22, 1.65, 9.39, 0.0, 0.88],
    # 4d
    "Zr": [2,  1.75, 1.33, 6.63, 2.0, 0.86],
    "Nb": [3,  1.64, 1.60, 6.76, 2.5, 0.78],
    "Mo": [4,  1.54, 2.16, 7.09, 2.5, 0.83],
    "Ru": [6,  1.46, 2.20, 7.36, 3.0, 0.82],
    "Rh": [7,  1.42, 2.28, 7.46, 3.0, 0.80],
    "Pd": [8,  1.39, 2.20, 8.34, 3.5, 0.78],
    "Ag": [9,  1.45, 1.93, 7.58, 6.0, 1.29],
    # 5d
    "Hf": [2,  1.75, 1.30, 6.83, 2.0, 0.85],
    "Ta": [3,  1.70, 1.50, 7.55, 2.5, 0.78],
    "W":  [4,  1.62, 2.36, 7.86, 2.5, 0.80],
    "Re": [5,  1.51, 1.90, 7.83, 3.0, 0.77],
    "Os": [6,  1.44, 2.20, 8.44, 3.0, 0.77],
    "Ir": [7,  1.41, 2.20, 8.97, 3.0, 0.77],
    "Pt": [8,  1.36, 2.28, 8.96, 3.5, 0.77],
    # Anions: [valence_e, r_cov, chi, r_ionic, polarizability]
    "O":  [6, 0.73, 3.44, 1.40, 0.80],
    "S":  [6, 1.04, 2.58, 1.84, 2.90],
    "Se": [6, 1.17, 2.55, 1.98, 3.77],
    "Te": [6, 1.37, 2.10, 2.21, 5.50],
}

SOC_LAMBDA = {
    "Ti": 20,  "V": 55,   "Cr": 60,  "Mn": 50,  "Fe": 50,  "Co": 88,
    "Ni": 100, "Cu": 110, "Zn": 0,
    "Zr": 100, "Nb": 150, "Mo": 200, "Ru": 300, "Rh": 350, "Pd": 400, "Ag": 200,
    "Hf": 400, "Ta": 600, "W": 800,  "Re": 1000, "Os": 1200, "Ir": 1500, "Pt": 1800,
}

STONER_I = {
    "Ti": 0.58, "V": 0.70, "Cr": 0.77, "Mn": 0.89, "Fe": 0.93, "Co": 0.99,
    "Ni": 1.01, "Cu": 0.73, "Zn": 0.0,
    "Zr": 0.33, "Nb": 0.38, "Mo": 0.44, "Ru": 0.50, "Rh": 0.53, "Pd": 0.56, "Ag": 0.38,
    "Hf": 0.30, "Ta": 0.33, "W": 0.38, "Re": 0.42, "Os": 0.46, "Ir": 0.48, "Pt": 0.50,
}

MADELUNG = {
    "NiAs": 1.560, "rock-salt": 1.748, "zincblende": 1.638,
    "wurtzite": 1.641, "MnP": 1.560, "pyrite": 1.800,
}

# d-band energy (eV) relative to anion p-band — for ZSA Delta
# Source: Bocquet et al. PRB 1996 / Zaanen et al. PRL 1985
D_BAND_ENERGY = {
    "Ti": -1.5, "V": -2.0, "Cr": -2.5, "Mn": -3.0, "Fe": -3.5,
    "Co": -4.0, "Ni": -4.5, "Cu": -5.5, "Zn": -7.0,
    "Zr": -1.0, "Nb": -1.5, "Mo": -2.0, "Ru": -3.0, "Rh": -3.5,
    "Pd": -4.0, "Ag": -5.0, "Hf": -0.8, "Ta": -1.2, "W": -1.8,
    "Re": -2.5, "Os": -3.0, "Ir": -3.5, "Pt": -4.0,
}

# Anion p-band energy (eV) — reference O = 0
ANION_P_ENERGY = {"O": 0.0, "S": 2.5, "Se": 3.5, "Te": 5.0}

TM_ELEMENTS = [
    "Ti", "V",  "Cr", "Mn", "Fe", "Co", "Ni", "Cu", "Zn",
    "Zr", "Nb", "Mo", "Ru", "Rh", "Pd", "Ag",
    "Hf", "Ta", "W",  "Re", "Os", "Ir", "Pt",
]

CHALCOGENS = ["O", "S", "Se", "Te"]

STOICHIOMETRIES = {
    "MC_NiAs":     {"structure": "NiAs",       "x": 1, "y": 1, "angle": 165.0},
    "MC_NaCl":     {"structure": "rock-salt",  "x": 1, "y": 1, "angle": 180.0},
    "MC_ZB":       {"structure": "zincblende", "x": 1, "y": 1, "angle": 109.5},
    "MC_WZ":       {"structure": "wurtzite",   "x": 1, "y": 1, "angle": 109.5},
    "MC_MnP":      {"structure": "MnP",        "x": 1, "y": 1, "angle": 155.0},
    "MC2_Pyrite":  {"structure": "pyrite",     "x": 1, "y": 2, "angle": 180.0},
    "M2C3_Corun":  {"structure": "NiAs",       "x": 2, "y": 3, "angle": 130.0},
    "M2C_AntiF":   {"structure": "zincblende", "x": 2, "y": 1, "angle": 109.5},
    "M3C4_Spinel": {"structure": "rock-salt",  "x": 3, "y": 4, "angle": 125.0},
}

# v6 feature list — 27 features (gap_log_proxy removed, 5 new added)
FEATURES_V6 = [
    # original 9
    "d_electrons", "Hubbard_U", "chi_diff", "bond_length_A",
    "bandwidth_W", "U_W_ratio", "bond_angle_deg", "ca_ratio",
    "exchange_corr_ratio",
    # v5 additions (10)
    "nd_half_fill", "r_ionic_ratio", "ionicity", "anion_polarizability",
    "soc_lambda", "crystal_field_10Dq", "jahn_teller_active",
    "superexchange_factor", "valence_electron_count", "madelung_energy",
    # binary flags (3)
    "is_oxide", "is_3d_metal", "is_5d_metal",
    # v6 new (5)
    "zsa_delta", "soc_sq", "stoner_split", "d_fill_ratio", "goldschmidt_t",
]

log.info("Atomic constants loaded. Feature set: %d features.", len(FEATURES_V6))
log.info("TM elements: %d | Chalcogens: %d | Stoichiometries: %d",
          len(TM_ELEMENTS), len(CHALCOGENS), len(STOICHIOMETRIES))

# =============================================================================
#  FEATURE ENGINEERING — v6.0 (27 features)
# =============================================================================

STRUCT_CA = {
    "NiAs": 1.534, "rock-salt": 1.000, "zincblende": 1.000,
    "wurtzite": 1.633, "MnP": 1.580, "pyrite": 1.000,
}

JT_CONFIGS = {1, 4, 7, 9}   # d^1, d^4, d^7, d^9 are JT-active


def compute_all_features_v6(M, C, x, y, structure, bond_angle_deg):
    """
    Compute the full 27-feature vector for compound M_x C_y in given structure.

    v6 additions vs v5.1:
      - zsa_delta    : Zaanen-Sawatzky-Allen charge-transfer energy
      - soc_sq       : lambda^2 (second-order SOC)
      - stoner_split : Stoner I x bandwidth W
      - d_fill_ratio : n_d / 10 (normalised d-filling)
      - goldschmidt_t: ionic-radius-based structure-stability proxy
      - gap_log_proxy REMOVED (circular with target)
    """
    ad_M = ATOMIC_DATA[M]   # [n_d, r_cov, chi, IE, U_Hubbard, r_ionic]
    ad_C = ATOMIC_DATA[C]   # [valence_e, r_cov, chi, r_ionic, polarizability]

    n_d   = ad_M[0];  r_cov_M = ad_M[1]; chi_M = ad_M[2]
    IE_M  = ad_M[3];  U       = ad_M[4]; r_ion_M = ad_M[5]

    val_C = ad_C[0];  r_cov_C = ad_C[1]; chi_C = ad_C[2]
    r_ion_C = ad_C[3]; pol_C = ad_C[4]

    # original features
    chi_diff = abs(chi_C - chi_M)

    # Refined bond length per stoichiometry
    if x == 1 and y == 1:
        bond_l = r_cov_M + r_cov_C
    elif x == 1 and y == 2:
        bond_l = r_cov_M + r_cov_C * 1.05
    elif x == 2 and y == 1:
        bond_l = r_cov_M * 1.05 + r_cov_C
    elif x == 2 and y == 3:
        bond_l = (r_cov_M + r_cov_C) * 1.02
    else:
        bond_l = r_cov_M + r_cov_C

    # Bandwidth W (Harrison model: W ~ 1/d^3.5)
    W = 3.5 / (bond_l ** 3.5)
    UW = U / (W + 1e-9)

    ca = STRUCT_CA.get(structure, 1.0)

    I_stoner = STONER_I.get(M, 0.5)
    exch_corr = I_stoner / (W + 1e-9)

    ba_deg = bond_angle_deg
    ba_rad = math.radians(ba_deg)

    # superexchange factor (Goodenough-Kanamori)
    sup_f = math.sin(ba_rad / 2) ** 2 * math.cos(ba_rad / 2) ** 2

    # valence electron count
    vec = x * n_d + y * val_C

    # Madelung energy
    mad_c = MADELUNG.get(structure, 1.6)
    z_eff = math.sqrt(abs(n_d - val_C))
    mad_e = mad_c * z_eff ** 2 / (r_ion_M + r_ion_C + 1e-9)

    # crystal field 10Dq (empirical: prop z_eff / bond^5)
    cf_10dq = 6.0 * z_eff / (bond_l ** 5 + 1e-9)

    # JT activity
    jt = int(n_d in JT_CONFIGS)

    # nd_half_fill: closeness to half-filling
    nd_hf = 1.0 - abs(n_d - 5.0) / 5.0

    # ionic radius ratio
    r_ratio = r_ion_M / (r_ion_C + 1e-9)

    # Phillips ionicity
    ionicity = (chi_C - chi_M) / (chi_C + chi_M + 1e-9)

    soc_lam = SOC_LAMBDA.get(M, 0.0)

    is_ox = int(C == "O")
    is_3d = int(M in ["Ti", "V", "Cr", "Mn", "Fe", "Co", "Ni", "Cu", "Zn"])
    is_5d = int(M in ["Hf", "Ta", "W", "Re", "Os", "Ir", "Pt"])

    # ── v6 new features ──────────────────────────────────────────────────
    # 1. ZSA charge-transfer energy: Delta = eps_d - eps_p - U/2
    eps_d = D_BAND_ENERGY.get(M, -3.0)
    eps_p = ANION_P_ENERGY.get(C, 0.0)
    zsa_d = eps_d - eps_p - U / 2.0

    # 2. SOC^2 (second-order perturbation -> orbital quenching)
    soc_sq = soc_lam ** 2 / 1e6

    # 3. Stoner exchange splitting E_ex = I x W
    stoner_split = I_stoner * W

    # 4. Normalised d-filling
    d_fill = n_d / 10.0

    # 5. Goldschmidt tolerance factor (ionic radii proxy for structure stability)
    r_ref = 1.4   # O2- reference ionic radius
    golds_t = (r_ion_M + r_ion_C) / (math.sqrt(2) * (r_ion_M + r_ref) + 1e-9)

    out = {
        "d_electrons":          float(n_d),
        "Hubbard_U":            float(U),
        "chi_diff":             float(chi_diff),
        "bond_length_A":        float(bond_l),
        "bandwidth_W":          float(W),
        "U_W_ratio":            float(UW),
        "bond_angle_deg":       float(ba_deg),
        "ca_ratio":             float(ca),
        "exchange_corr_ratio":  float(exch_corr),
        "nd_half_fill":         float(nd_hf),
        "r_ionic_ratio":        float(r_ratio),
        "ionicity":             float(ionicity),
        "anion_polarizability": float(pol_C),
        "soc_lambda":           float(soc_lam),
        "crystal_field_10Dq":   float(cf_10dq),
        "jahn_teller_active":   float(jt),
        "superexchange_factor": float(sup_f),
        "valence_electron_count": float(vec),
        "madelung_energy":      float(mad_e),
        "is_oxide":             float(is_ox),
        "is_3d_metal":          float(is_3d),
        "is_5d_metal":          float(is_5d),
        # v6 new
        "zsa_delta":            float(zsa_d),
        "soc_sq":               float(soc_sq),
        "stoner_split":         float(stoner_split),
        "d_fill_ratio":         float(d_fill),
        "goldschmidt_t":        float(golds_t),
        # internal (not in FEATURES_V6, used for other calcs)
        "_bl": bond_l, "_W": W, "_UW": UW, "_chi_diff": chi_diff, "_ca": ca,
    }
    return out


log.info("compute_all_features_v6() defined (%d features).", len(FEATURES_V6))

# =============================================================================
#  PDF LITERATURE MINING — v6.0 SCRIPT EDITION
#  Upgraded for large libraries (1000+ papers):
#    - Parallel extraction (ThreadPoolExecutor) instead of sequential
#    - Progress reporting every PDF_BATCH_SIZE files (no tqdm dependency)
#    - Per-file timeout to avoid one corrupt/huge PDF stalling the whole run
#    - Skips files above PDF_SKIP_LARGER_THAN_MB (likely scanned image PDFs
#      that pypdf/pdftotext cannot usefully extract text from anyway)
#    - Smarter formula/number matching: requires the formula and numbers to
#      be reasonably close together (same line OR formula + next line) to
#      reduce false positives at this scale
#    - Writes a CSV of all candidate rows to OUTPUT_DIR for offline review,
#      instead of only printing to console (which is impractical for 1000+
#      papers worth of candidates)
# =============================================================================

_FORMULA_PAT = re.compile(r'\b([A-Z][a-z]?)(\d{0,2})([A-Z][a-z]?)(\d{0,2})\b')
_NUM_PAT = re.compile(r'-?\d+\.\d+|-?\d+')

_KNOWN_METALS = {
    "Ti", "V", "Cr", "Mn", "Fe", "Co", "Ni", "Cu", "Zn",
    "Zr", "Nb", "Mo", "Ru", "Rh", "Pd", "Ag",
    "Hf", "Ta", "W", "Re", "Os", "Ir", "Pt",
}
_KNOWN_CHALCOGENS = {"O", "S", "Se", "Te"}


def _maybe_mount_drive(directory: str) -> bool:
    """
    Mount Google Drive only if running inside Colab AND the path lives
    under /content/drive. For a standalone script (this file), this is a
    no-op in 99% of cases — Drive access for scripts is normally handled
    via rclone, the Google Drive desktop sync client, or a service-account
    mounted filesystem, all of which already expose a normal local path.
    """
    if not str(directory).startswith("/content/drive"):
        return True
    try:
        from google.colab import drive  # type: ignore
        drive.mount("/content/drive", force_remount=False)
        log.info("Google Drive mounted at /content/drive")
        return True
    except Exception as e:
        log.warning("Drive mount failed (expected outside Colab): %s", e)
        return False


def _read_pdf_text(pdf_path: Path, timeout: int = PDF_TIMEOUT_SEC) -> str:
    """
    Extract plain text from *pdf_path*.
    Strategy:
      1. pdftotext -layout  (poppler — best column alignment, if installed)
      2. pypdf              (pure-Python fallback, always available)
    Returns an empty string on failure or timeout.
    """
    path_str = str(pdf_path)

    # attempt 1: pdftotext (poppler)
    try:
        result = subprocess.run(
            ["pdftotext", "-layout", path_str, "-"],
            capture_output=True, text=True, timeout=timeout,
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout
    except FileNotFoundError:
        pass  # poppler not installed — fall through to pypdf
    except subprocess.TimeoutExpired:
        log.warning("pdftotext timeout (%ds): %s", timeout, pdf_path.name)
    except Exception as e:
        log.warning("pdftotext error on %s: %s", pdf_path.name, e)

    # attempt 2: pypdf (pure Python, no external binary needed)
    try:
        import pypdf
    except ImportError:
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "pypdf", "-q"]
            )
            import pypdf
        except Exception as e:
            log.warning("pypdf install failed: %s", e)
            return ""
    try:
        reader = pypdf.PdfReader(path_str)
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception as e:
        log.warning("pypdf failed on %s: %s", pdf_path.name, e)
        return ""


def _extract_records_from_text(text: str, source: str) -> list[dict]:
    """
    Scan extracted text for (compound, band_gap_eV, magnetization_muB)
    candidate triples.

    v6.0-script smarter matching vs the notebook's per-line-only scanner:
    if the formula line itself has fewer than 2 trailing numbers, also
    check the *next* line — many papers put a formula in a table-row
    label cell and the numeric values in the following wrapped line.
    """
    records = []
    lines = [ln.strip() for ln in text.splitlines()]

    for i, line in enumerate(lines):
        if not line:
            continue

        fm = _FORMULA_PAT.search(line)
        if not fm:
            continue

        M_cand, C_cand = fm.group(1), fm.group(3)
        if M_cand not in _KNOWN_METALS or C_cand not in _KNOWN_CHALCOGENS:
            continue

        # Numbers on the same line, after the formula
        rest = line[fm.end():]
        nums = [float(n) for n in _NUM_PAT.findall(rest)]

        # Smarter fallback: check the next line too (wrapped table rows)
        used_next_line = False
        if len(nums) < 2 and i + 1 < len(lines):
            nums_next = [float(n) for n in _NUM_PAT.findall(lines[i + 1])]
            if len(nums_next) >= 2:
                nums = nums_next
                used_next_line = True

        if len(nums) < 2:
            continue

        gap, mag = nums[0], nums[1]

        # Physical plausibility gate
        if not (0.0 <= gap <= 12.0) or not (0.0 <= mag <= 20.0):
            continue

        x_str = fm.group(2) or "1"
        y_str = fm.group(4) or "1"
        compound = (
            f"{M_cand}{'' if x_str == '1' else x_str}"
            f"{C_cand}{'' if y_str == '1' else y_str}"
        )

        records.append({
            "compound": compound,
            "M": M_cand,
            "C": C_cand,
            "band_gap_eV": round(gap, 4),
            "magnetization_muB": round(mag, 4),
            "raw_line": (line + " | next: " + lines[i + 1]) if used_next_line else line,
            "source": source,
        })

    return records


def extract_tables_from_pdf(pdf_path: Path) -> list[dict]:
    """Single-PDF extraction entry point (used by the thread pool below)."""
    pdf_path = Path(pdf_path)
    text = _read_pdf_text(pdf_path)
    if not text:
        return []
    return _extract_records_from_text(text, source=pdf_path.stem)


def _scan_one_pdf(pdf_path: Path) -> tuple[Path, list[dict], str | None]:
    """Wrapper for parallel execution: returns (path, records, error)."""
    try:
        size_mb = pdf_path.stat().st_size / (1024 * 1024)
        if size_mb > PDF_SKIP_LARGER_THAN_MB:
            return pdf_path, [], f"skipped (size {size_mb:.1f} MB > limit)"
        recs = extract_tables_from_pdf(pdf_path)
        return pdf_path, recs, None
    except Exception as e:
        return pdf_path, [], str(e)


def scan_paper_directory(paper_dir: str) -> list[dict]:
    """
    Scan *paper_dir* for *.pdf files (recursively) and extract compound data
    from each one, in parallel, with progress reporting suited to libraries
    of 1000+ papers (e.g. a Google Drive sync folder).

    Writes two files to OUTPUT_DIR:
      - pdf_mining_candidates_v6.csv : every candidate row found, for review
      - pdf_mining_summary_v6.json   : per-file record counts + errors

    Returns the flat list of candidate record dicts.
    """
    dir_path = Path(paper_dir)
    if not dir_path.exists():
        log.warning("PDF directory not found: %s", dir_path)
        return []
    if not dir_path.is_dir():
        log.warning("Path is not a directory: %s", dir_path)
        return []

    # Recursive glob — Google Drive sync folders are often organised into
    # subfolders (e.g. by the pdf_organizer.py classification categories)
    pdf_files = sorted(dir_path.rglob("*.pdf"))
    if not pdf_files:
        log.warning("No PDF files found under: %s", dir_path)
        return []

    n_total = len(pdf_files)
    log.info("Found %d PDF file(s) under %s (recursive scan)", n_total, dir_path)
    log.info("Mining with %d parallel workers, batch report every %d files",
              PDF_MAX_WORKERS, PDF_BATCH_SIZE)

    all_records: list[dict] = []
    errors: list[dict] = []
    t_start = time.time()
    n_done = 0
    n_with_hits = 0

    with ThreadPoolExecutor(max_workers=PDF_MAX_WORKERS) as executor:
        futures = {executor.submit(_scan_one_pdf, p): p for p in pdf_files}

        for future in as_completed(futures):
            pdf_path, recs, err = future.result()
            n_done += 1

            if err:
                errors.append({"file": str(pdf_path), "error": err})
            elif recs:
                n_with_hits += 1
                all_records.extend(recs)

            if n_done % PDF_BATCH_SIZE == 0 or n_done == n_total:
                elapsed = time.time() - t_start
                rate = n_done / elapsed if elapsed > 0 else 0.0
                eta_sec = (n_total - n_done) / rate if rate > 0 else 0.0
                log.info(
                    "  [%d/%d]  %.1f%%  |  %d candidate rows so far  |  "
                    "%.1f files/s  |  ETA %.0fs",
                    n_done, n_total, 100.0 * n_done / n_total,
                    len(all_records), rate, eta_sec,
                )

    elapsed_total = time.time() - t_start
    log.info("PDF mining complete: %d files in %.1fs (%.1f files/s)",
              n_total, elapsed_total, n_total / max(elapsed_total, 1e-9))
    log.info("  Files with >=1 candidate row: %d / %d", n_with_hits, n_total)
    log.info("  Total candidate rows extracted: %d", len(all_records))
    if errors:
        log.warning("  Files with errors/skips: %d", len(errors))

    # Save candidates to CSV for offline review (essential at this scale —
    # printing 1000+ papers' worth of rows to console is not usable)
    if all_records:
        cand_df = pd.DataFrame(all_records)
        cand_path = os.path.join(OUTPUT_DIR, "pdf_mining_candidates_v6.csv")
        cand_df.to_csv(cand_path, index=False)
        log.info("  Candidate rows written to: %s", cand_path)
        log.info("  Review this file, then copy confirmed rows into "
                  "PAPER_DB_CSV_PATH or PAPER_DB_MANUAL for the next run.")

    summary = {
        "directory": str(dir_path),
        "total_pdfs": n_total,
        "files_with_hits": n_with_hits,
        "total_candidate_rows": len(all_records),
        "elapsed_seconds": round(elapsed_total, 1),
        "errors": errors[:200],  # cap stored errors to keep file small
    }
    summary_path = os.path.join(OUTPUT_DIR, "pdf_mining_summary_v6.json")
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)
    log.info("  Mining summary written to: %s", summary_path)

    return all_records


def run_pdf_mining_stage() -> list[dict]:
    """Top-level entry point called from main()."""
    if not PAPER_DIR:
        log.info("PAPER_DIR not set — skipping PDF mining stage. "
                  "Set MXCY_PAPER_DIR env var or PAPER_DIR constant to enable.")
        return []
    _maybe_mount_drive(PAPER_DIR)
    return scan_paper_directory(PAPER_DIR)

# =============================================================================
#  PAPER DATABASE INTEGRATION
# =============================================================================

PAPER_DB_MANUAL: list[dict] = [
    # Paste confirmed rows here, e.g. from Reggad et al., Physica B 526 (2017):
    # {"compound": "FeS2", "M": "Fe", "C": "S", "structure": "MC2_Pyrite",
    #  "band_gap_eV": 0.95, "magnetization_muB": 0.0, "source": "Reggad2017"},
]


def load_paper_db() -> dict:
    """
    Load paper-extracted compound data from (in order of precedence):
      1. PAPER_DB_MANUAL (hand-curated, highest trust)
      2. PAPER_DB_CSV_PATH (your reviewed pdf_mining_candidates_v6.csv subset)

    Values from papers supplement Materials Project data — they only fill
    gaps where MP has no entry for that exact (M, C, structure) combination.
    """
    records = list(PAPER_DB_MANUAL)

    if PAPER_DB_CSV_PATH and os.path.isfile(PAPER_DB_CSV_PATH):
        try:
            df_p = pd.read_csv(PAPER_DB_CSV_PATH)
            required = {"compound", "M", "C", "structure",
                        "band_gap_eV", "magnetization_muB"}
            if required.issubset(df_p.columns):
                records += df_p.to_dict("records")
                log.info("Loaded %d rows from CSV: %s", len(df_p), PAPER_DB_CSV_PATH)
            else:
                missing = required - set(df_p.columns)
                log.warning("CSV missing required columns: %s", missing)
        except Exception as e:
            log.warning("Could not read CSV %s: %s", PAPER_DB_CSV_PATH, e)

    paper_dict = {}
    for r in records:
        try:
            key = (r["M"], r["C"], r["structure"])
            paper_dict[key] = {
                "band_gap_eV": float(r["band_gap_eV"]),
                "magnetization_muB": float(r["magnetization_muB"]),
                "source": r.get("source", "paper"),
            }
        except (KeyError, ValueError) as e:
            log.warning("Skipping malformed paper DB row %s: %s", r, e)

    log.info("Paper DB: %d compound entries loaded.", len(paper_dict))
    return paper_dict


# =============================================================================
#  MATERIALS PROJECT FETCH
# =============================================================================

MAX_HULL = 0.20  # eV/atom
ORDERING_MAP = {"NM": 0, "AFM": 1, "FM": 2}


def fetch_real_data_v6(api_key: str, paper_db: dict | None = None) -> dict:
    """
    Fetch MxCy compounds from Materials Project and merge with paper_db.
    Paper DB values supplement MP where MP has no entry (MP-DFT is always
    the preferred reference where both exist).
    """
    if not api_key or len(api_key) < 10:
        log.warning("No valid MP_API_KEY set — returning empty dict. "
                    "Set the MP_API_KEY environment variable to enable fetching.")
        return {}

    try:
        from mp_api.client import MPRester
    except ImportError:
        log.error("mp_api not installed. Run: pip install mp-api")
        return {}

    real_dict: dict = {}
    log.info("Fetching from Materials Project ...")

    with MPRester(api_key) as mpr:
        for M in TM_ELEMENTS:
            for C in CHALCOGENS:
                try:
                    results = mpr.materials.summary.search(
                        elements=[M, C],
                        num_elements=[2],
                        energy_above_hull=(0.0, MAX_HULL),
                        fields=[
                            "formula_pretty", "band_gap", "ordering",
                            "energy_above_hull", "formation_energy_per_atom",
                            "total_magnetization", "structure",
                        ],
                    )
                except Exception as e:
                    log.warning("MP query failed for %s-%s: %s", M, C, e)
                    continue

                for r in results:
                    formula = r.formula_pretty
                    gap = float(r.band_gap or 0.0)
                    hull = float(r.energy_above_hull or 0.0)
                    fe = float(r.formation_energy_per_atom or float("nan"))
                    mag = abs(float(r.total_magnetization or 0.0))

                    ord_str = str(r.ordering).upper() if r.ordering else "NM"
                    if "AFM" in ord_str:
                        ord_label = 1
                    elif "FM" in ord_str or "FIM" in ord_str:
                        ord_label = 2
                    else:
                        ord_label = 0

                    for stoich_key, st in STOICHIOMETRIES.items():
                        x_s, y_s = st["x"], st["y"]
                        expected = (
                            f"{M}{C}" if x_s == 1 and y_s == 1 else
                            f"{M}{C}2" if x_s == 1 and y_s == 2 else
                            f"{M}2{C}3" if x_s == 2 and y_s == 3 else
                            f"{M}2{C}" if x_s == 2 and y_s == 1 else
                            f"{M}3{C}4" if x_s == 3 and y_s == 4 else None
                        )
                        if expected and expected == formula:
                            key = (M, C, stoich_key)
                            real_dict[key] = {
                                "band_gap_eV": gap,
                                "magnetization_muB": mag,
                                "hull_eV_per_atom": hull,
                                "formation_energy_per_atom": fe,
                                "ordering_label": ord_label,
                                "ordering_str": ["NM", "AFM", "FM"][ord_label],
                                "source": "MP",
                            }

    log.info("MP: %d entries fetched.", len(real_dict))

    if paper_db:
        merged = 0
        for key, val in paper_db.items():
            if key not in real_dict:
                real_dict[key] = {
                    "band_gap_eV": val["band_gap_eV"],
                    "magnetization_muB": val["magnetization_muB"],
                    "hull_eV_per_atom": 0.0,
                    "formation_energy_per_atom": float("nan"),
                    "ordering_label": 2 if val["magnetization_muB"] > 0.5 else 0,
                    "ordering_str": "FM" if val["magnetization_muB"] > 0.5 else "NM",
                    "source": val.get("source", "paper"),
                }
                merged += 1
        log.info("Paper DB: %d additional entries merged.", merged)

    log.info("Total real entries: %d", len(real_dict))
    return real_dict


# =============================================================================
#  EMPIRICAL GAP & MAGNETIZATION  (for non-MP, non-paper compounds)
# =============================================================================

def predict_half_metal_empirical(M, C, structure, W):
    ad_M = ATOMIC_DATA[M]; ad_C = ATOMIC_DATA[C]
    U = ad_M[4]; chi_M = ad_M[2]; chi_C = ad_C[2]
    UW = U / (W + 1e-9)
    chi_diff = abs(chi_C - chi_M)
    spin_pol = (chi_diff + UW) * 10
    return (UW > 10 or chi_diff > 1.5), min(spin_pol, 100.0)


def calc_empirical_band_gap(M, C, UW, chi_diff, n_total, structure, ca):
    base = 0.8 * UW + 0.5 * chi_diff - 0.3 * n_total + 0.2 * ca
    return max(0.0, round(base, 3))


def calc_empirical_magnetization(M, C, x, UW, W, structure):
    I = STONER_I.get(M, 0.5)
    if I * W > 1.0:
        nd = ATOMIC_DATA[M][0]
        return round(min(nd, 10 - nd) * 1.0, 2)
    return 0.0


def build_full_dataset_v6(real_dict: dict) -> pd.DataFrame:
    """Build the full combinatorial MxCy x stoichiometry grid with real data
    where available and physics-based empirical estimates elsewhere."""
    log.info("Building v6.0 dataset (%d features) ...", len(FEATURES_V6))
    rows = []
    for M in TM_ELEMENTS:
        for C in CHALCOGENS:
            for stoich_key, st in STOICHIOMETRIES.items():
                structure = st["structure"]
                x, y, angle = st["x"], st["y"], st["angle"]

                feats = compute_all_features_v6(M, C, x, y, structure, angle)
                bl, W, UW = feats["_bl"], feats["_W"], feats["_UW"]
                chi_diff, ca = feats["_chi_diff"], feats["_ca"]
                is_hm, spin_pol = predict_half_metal_empirical(M, C, structure, W)

                combo_key = (M, C, stoich_key)
                if combo_key in real_dict:
                    rd = real_dict[combo_key]
                    gap = rd["band_gap_eV"]; mag = rd["magnetization_muB"]
                    hull = rd["hull_eV_per_atom"]; fe = rd["formation_energy_per_atom"]
                    ordering_label = rd.get("ordering_label", 0)
                    ordering_str = rd.get("ordering_str", "NM")
                    is_real = True
                    is_hm_r = (gap < 0.01 and mag > 0.5) or is_hm
                    sp_r = min(100., 100. * mag / (mag + W + 1e-6)) if gap < 0.01 and mag > 0.5 else spin_pol
                else:
                    gap = calc_empirical_band_gap(M, C, UW, chi_diff, x + y, structure, ca)
                    mag = calc_empirical_magnetization(M, C, x, UW, W, structure)
                    hull = float("nan"); fe = float("nan")
                    ordering_label = -1; ordering_str = "Unknown"
                    is_real = False; is_hm_r = is_hm; sp_r = spin_pol

                formula = (
                    f"{M}{C}" if x == 1 and y == 1 else
                    f"{M}{C}2" if x == 1 and y == 2 else
                    f"{M}2{C}3" if x == 2 and y == 3 else
                    f"{M}2{C}" if x == 2 and y == 1 else
                    f"{M}3{C}4" if x == 3 and y == 4 else f"{M}{x}{C}{y}"
                )

                rows.append({
                    "Compound": formula, "M": M, "C": C, "Stoichiometry": stoich_key,
                    **{f: feats[f] for f in FEATURES_V6},
                    "band_gap_eV": gap, "magnetization_muB": mag,
                    "Is_Half_Metal": bool(is_hm_r), "Spin_Polarization": sp_r,
                    "hull_eV_per_atom": hull, "formation_energy_per_atom": fe,
                    "Is_Real": is_real, "ordering_label": ordering_label,
                    "ordering_str": ordering_str,
                })

    df = pd.DataFrame(rows)
    n_real = df["Is_Real"].sum()
    log.info("Total compounds: %d | Real (MP+papers): %d", len(df), n_real)
    n_nm = (df[df.Is_Real]["ordering_label"] == 0).sum()
    n_afm = (df[df.Is_Real]["ordering_label"] == 1).sum()
    n_fm = (df[df.Is_Real]["ordering_label"] == 2).sum()
    log.info("Real ordering: NM=%d AFM=%d FM=%d", n_nm, n_afm, n_fm)
    return df

# =============================================================================
#  ML TRAINING — v6.0
#  Domain-split gap ensembles, ordering-aware magnetisation, bootstrap CI.
# =============================================================================

def train_ml_v6(df: pd.DataFrame, n_bootstrap: int = 50, ci_alpha: float = 0.90):
    """
    Train the v6.0 ML pipeline.

    Returns
    -------
    imp        : DataFrame — feature importances
    cv_results : dict — CV + test metrics
    model_path : str — path to saved tuned model
    df_te      : DataFrame — test-set predictions
    """
    log.info("Training v6.0 ML model ...")
    df_real = df[df["Is_Real"]].copy().reset_index(drop=True)
    n_real = len(df_real)
    log.info("Real MP+paper samples: %d", n_real)
    if n_real < 60:
        log.warning("Need >= 60 real entries to train. Skipping ML stage.")
        return pd.DataFrame(), {}, None, df_real.head(0)

    FEAT = FEATURES_V6  # 27 features

    # ── Stratified train/test split ──────────────────────────────────────
    strat = df_real["ordering_label"].values
    n_test = max(50, int(0.20 * n_real))
    sss = StratifiedShuffleSplit(n_splits=1, test_size=n_test, random_state=42)
    tr_i, te_i = next(sss.split(df_real, strat))
    df_tr = df_real.iloc[tr_i].reset_index(drop=True)
    df_te = df_real.iloc[te_i].reset_index(drop=True)
    n_tr = len(df_tr)
    log.info("Train: %d | Test: %d (stratified by ordering)", n_tr, n_test)

    X_tr = df_tr[FEAT].values; X_te = df_te[FEAT].values
    g_tr = df_tr["band_gap_eV"].values; g_te = df_te["band_gap_eV"].values
    m_tr = df_tr["magnetization_muB"].values; m_te = df_te["magnetization_muB"].values
    o_tr = df_tr["ordering_label"].values; o_te = df_te["ordering_label"].values

    # Augment features with ordering (for magnetisation model)
    X_tr_mag = np.column_stack([X_tr, o_tr.reshape(-1, 1)])
    X_te_mag = np.column_stack([X_te, o_te.reshape(-1, 1)])

    log.info("[Train] NM:%d | AFM:%d | FM:%d",
              (o_tr == 0).sum(), (o_tr == 1).sum(), (o_tr == 2).sum())

    # ════════════════════════════════════════════════════════════════════
    #  ORDERING CLASSIFIER (trained first — needed for test-time mag)
    # ════════════════════════════════════════════════════════════════════
    clf_ord = None
    if len(np.unique(o_tr)) >= 2:
        params_ord = {"n_estimators": [300, 500], "max_features": ["sqrt", "log2"],
                       "min_samples_leaf": [1, 2]}
        cv_ord = min(5, max(2, n_tr // 4))
        sr = RandomizedSearchCV(
            RandomForestClassifier(random_state=42, n_jobs=-1, class_weight="balanced"),
            params_ord, n_iter=8, cv=cv_ord, scoring="balanced_accuracy",
            random_state=42, n_jobs=-1)
        sr.fit(X_tr, o_tr)
        clf_ord = sr.best_estimator_
        joblib.dump(clf_ord, os.path.join(OUTPUT_DIR, "mxcy_ordering_clf_v6.pkl"))
        ord_acc_tr = accuracy_score(o_te, clf_ord.predict(X_te))
        log.info("Ordering classifier | test acc: %.3f", ord_acc_tr)
    else:
        ord_acc_tr = float("nan")

    # ════════════════════════════════════════════════════════════════════
    #  BAND GAP — log1p 3-model ensemble (RF + HGB + XGB)
    # ════════════════════════════════════════════════════════════════════
    ins_mask = g_tr >= 0.01
    g_tr_log = np.log1p(g_tr)
    log.info("[Gap] insulators in train: %d  metals: %d",
              ins_mask.sum(), (~ins_mask).sum())

    clf_metal = None
    if ins_mask.sum() >= 5 and (~ins_mask).sum() >= 5:
        clf_metal = Pipeline([
            ("sc", StandardScaler()),
            ("rf", RandomForestClassifier(n_estimators=500, class_weight="balanced",
                                           random_state=42, n_jobs=-1))
        ])
        clf_metal.fit(X_tr, ins_mask.astype(int))
        joblib.dump(clf_metal, os.path.join(OUTPUT_DIR, "mxcy_metal_clf_v6.pkl"))

    def _fit_gap_ensemble(X, y_log, tag=""):
        n_cv = min(5, max(2, len(X) // 4))
        if len(X) < 8:
            rf = RandomForestRegressor(n_estimators=300, random_state=42, n_jobs=-1)
            rf.fit(X, y_log)
            return {"rf": rf, "hgb": None, "xgb": None}
        params_rf = {"n_estimators": [300, 500], "max_features": ["sqrt", "log2", 0.6],
                     "min_samples_leaf": [1, 2, 3], "max_depth": [None, 10, 20]}
        s_rf = RandomizedSearchCV(
            RandomForestRegressor(random_state=42, n_jobs=-1),
            params_rf, n_iter=12, cv=n_cv, scoring="r2", random_state=42, n_jobs=-1)
        s_rf.fit(X, y_log)
        hgb = HistGradientBoostingRegressor(max_iter=400, learning_rate=0.04,
                                             max_depth=5, random_state=42)
        hgb.fit(X, y_log)
        xgb_m = xgb.XGBRegressor(n_estimators=400, learning_rate=0.04, max_depth=5,
                                  subsample=0.8, colsample_bytree=0.8,
                                  random_state=42, n_jobs=-1, verbosity=0)
        xgb_m.fit(X, y_log)
        log.info("  %s RF best params: %s", tag, s_rf.best_params_)
        return {"rf": s_rf.best_estimator_, "hgb": hgb, "xgb": xgb_m}

    gap_models = {}
    if ins_mask.sum() >= 6:
        oxide_mask = (df_tr["C"] == "O").values
        ins_ox = ins_mask & oxide_mask
        ins_nox = ins_mask & ~oxide_mask
        log.info("[Gap] oxide insulators: %d  non-oxide insulators: %d",
                  ins_ox.sum(), ins_nox.sum())
        if ins_ox.sum() >= 6:
            gap_models["ox"] = _fit_gap_ensemble(X_tr[ins_ox], g_tr_log[ins_ox], "ox")
        if ins_nox.sum() >= 6:
            gap_models["nox"] = _fit_gap_ensemble(X_tr[ins_nox], g_tr_log[ins_nox], "nox")
        gap_models["all"] = _fit_gap_ensemble(X_tr[ins_mask], g_tr_log[ins_mask], "all")

    iso_gap = None
    if "all" in gap_models and ins_mask.sum() >= 10:
        rf_all = gap_models["all"]["rf"]
        raw_tr = np.expm1(rf_all.predict(X_tr[ins_mask]))
        iso_gap = IsotonicRegression(out_of_bounds="clip")
        iso_gap.fit(raw_tr, g_tr[ins_mask])

    def _predict_gap(X, is_oxide_arr):
        n = len(X)
        preds = np.zeros(n)
        is_ins = (clf_metal.predict(X).astype(bool) if clf_metal is not None
                  else np.ones(n, bool))
        for idx in np.where(is_ins)[0]:
            x_i = X[idx:idx + 1]
            p = []
            key = "ox" if is_oxide_arr[idx] else "nox"
            for k in [key, "all"]:
                if k in gap_models:
                    gm = gap_models[k]
                    p.append(np.expm1(gm["rf"].predict(x_i)[0]))
                    if gm["hgb"]:
                        p.append(np.expm1(gm["hgb"].predict(x_i)[0]))
                    if gm["xgb"]:
                        p.append(np.expm1(gm["xgb"].predict(x_i)[0]))
            raw = float(np.mean(p)) if p else 0.0
            preds[idx] = iso_gap.transform([raw])[0] if iso_gap else raw
        return np.clip(preds, 0, None)

    # ════════════════════════════════════════════════════════════════════
    #  MAGNETIZATION — v6 fix: ordering passed in as a feature
    # ════════════════════════════════════════════════════════════════════
    is_mag_tr = (m_tr > 0.5).astype(int)
    _n_pos = is_mag_tr.sum(); _n_neg = len(is_mag_tr) - _n_pos
    cw_nm = {0: 1.0, 1: min(3.0, float(_n_neg) / max(1, _n_pos))}
    clf_nm_mag = Pipeline([
        ("sc", StandardScaler()),
        ("rf", RandomForestClassifier(n_estimators=500, class_weight=cw_nm,
                                       random_state=42, n_jobs=-1))
    ])
    clf_nm_mag.fit(X_tr, is_mag_tr)
    joblib.dump(clf_nm_mag, os.path.join(OUTPUT_DIR, "mxcy_nm_mag_clf_v6.pkl"))
    log.info("NM/magnetic classifier (NM:%d | mag:%d)",
              (is_mag_tr == 0).sum(), _n_pos)

    mag_mask_tr = is_mag_tr.astype(bool)
    rf_mag = None; xgb_mag = None; hgb_mag = None
    if mag_mask_tr.sum() >= 10:
        X_mag_sub = X_tr_mag[mag_mask_tr]
        y_mag_sub = m_tr[mag_mask_tr]
        n_cv_m = min(5, max(2, mag_mask_tr.sum() // 4))
        params_m = {"n_estimators": [300, 500], "max_features": ["sqrt", "log2", 0.7],
                    "min_samples_leaf": [1, 2], "max_depth": [None, 12, 24]}
        sm = RandomizedSearchCV(
            RandomForestRegressor(random_state=42, n_jobs=-1),
            params_m, n_iter=12, cv=n_cv_m, scoring="r2",
            random_state=42, n_jobs=-1)
        sm.fit(X_mag_sub, y_mag_sub)
        rf_mag = sm.best_estimator_
        log.info("Best RF mag params: %s", sm.best_params_)
        xgb_mag = xgb.XGBRegressor(n_estimators=400, learning_rate=0.04, max_depth=5,
                                    subsample=0.8, colsample_bytree=0.8,
                                    random_state=42, n_jobs=-1, verbosity=0)
        xgb_mag.fit(X_mag_sub, y_mag_sub)
        hgb_mag = HistGradientBoostingRegressor(max_iter=400, learning_rate=0.04,
                                                 max_depth=5, random_state=42)
        hgb_mag.fit(X_mag_sub, y_mag_sub)
    elif mag_mask_tr.sum() >= 5:
        X_mag_sub = X_tr_mag[mag_mask_tr]
        y_mag_sub = m_tr[mag_mask_tr]
        rf_mag = RandomForestRegressor(n_estimators=300, random_state=42, n_jobs=-1)
        rf_mag.fit(X_mag_sub, y_mag_sub)

    def _predict_mag(X, X_with_ord):
        n = len(X)
        preds = np.zeros(n)
        proba = (clf_nm_mag.predict_proba(X)[:, 1]
                 if hasattr(clf_nm_mag, "predict_proba") else
                 clf_nm_mag.predict(X).astype(float))
        mag_flag = proba >= 0.45
        mag_idx = np.where(mag_flag)[0]
        if mag_idx.size > 0 and rf_mag is not None:
            p = [np.clip(rf_mag.predict(X_with_ord[mag_idx]), 0, None)]
            if xgb_mag:
                p.append(np.clip(xgb_mag.predict(X_with_ord[mag_idx]), 0, None))
            if hgb_mag:
                p.append(np.clip(hgb_mag.predict(X_with_ord[mag_idx]), 0, None))
            preds[mag_idx] = np.mean(p, axis=0)
        return preds

    # ════════════════════════════════════════════════════════════════════
    #  TEST-SET EVALUATION
    # ════════════════════════════════════════════════════════════════════
    log.info("=== Held-out test evaluation (%d compounds) ===", n_test)
    ox_te_arr = (df_te["C"] == "O").values
    gap_te_pred = _predict_gap(X_te, ox_te_arr)
    o_te_pred = clf_ord.predict(X_te) if clf_ord is not None else o_te
    X_te_mag_pred = np.column_stack([X_te, o_te_pred.reshape(-1, 1)])
    mag_te_pred = _predict_mag(X_te, X_te_mag_pred)

    r2_gap = r2_score(g_te, gap_te_pred)
    mae_gap = mean_absolute_error(g_te, gap_te_pred)
    r2_mag = r2_score(m_te, mag_te_pred)
    mae_mag = mean_absolute_error(m_te, mag_te_pred)
    ord_acc = (accuracy_score(o_te, clf_ord.predict(X_te))
               if clf_ord is not None else float("nan"))

    log.info("R2(gap)=%.4f | MAE(gap)=%.4f eV", r2_gap, mae_gap)
    log.info("R2(mag)=%.4f | MAE(mag)=%.4f muB", r2_mag, mae_mag)
    log.info("Ordering accuracy = %.4f", ord_acc)

    # ════════════════════════════════════════════════════════════════════
    #  STRATIFIED 5-FOLD CV
    # ════════════════════════════════════════════════════════════════════
    log.info("Running 5-fold stratified CV ...")
    skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
    cv_gap_r2, cv_mag_r2 = [], []

    for fold_i, (cv_tr_i, cv_te_i) in enumerate(skf.split(X_tr, o_tr)):
        Xc_tr, Xc_te = X_tr[cv_tr_i], X_tr[cv_te_i]
        gc_tr, gc_te = g_tr[cv_tr_i], g_tr[cv_te_i]
        mc_tr, mc_te = m_tr[cv_tr_i], m_tr[cv_te_i]
        oc_tr, oc_te = o_tr[cv_tr_i], o_tr[cv_te_i]

        if len(np.unique(oc_tr)) >= 2:
            clf_ord_cv = RandomForestClassifier(n_estimators=200, class_weight="balanced",
                                                 random_state=42)
            clf_ord_cv.fit(Xc_tr, oc_tr)
            oc_te_pred_cv = clf_ord_cv.predict(Xc_te)
        else:
            oc_te_pred_cv = oc_te

        ins_c = gc_tr >= 0.01
        g_log_c = np.log1p(gc_tr)
        yp_g = np.zeros(len(cv_te_i))
        if ins_c.sum() >= 3:
            mc_g = xgb.XGBRegressor(n_estimators=200, learning_rate=0.05, max_depth=4,
                                     random_state=42, n_jobs=-1, verbosity=0)
            mc_g.fit(Xc_tr[ins_c], g_log_c[ins_c])
            if clf_metal is not None:
                mc_clf = RandomForestClassifier(n_estimators=200, random_state=42,
                                                 class_weight="balanced")
                if (~ins_c).sum() >= 2:
                    mc_clf.fit(Xc_tr, ins_c.astype(int))
                    is_ins_cv = mc_clf.predict(Xc_te).astype(bool)
                else:
                    is_ins_cv = np.ones(len(cv_te_i), bool)
                if is_ins_cv.sum() > 0:
                    yp_g[is_ins_cv] = np.expm1(mc_g.predict(Xc_te[is_ins_cv]))
            else:
                yp_g = np.expm1(mc_g.predict(Xc_te))
        cv_gap_r2.append(r2_score(gc_te, yp_g))

        is_mag_c = (mc_tr > 0.5).astype(int)
        mag_clf_c = RandomForestClassifier(n_estimators=200, class_weight="balanced",
                                            random_state=42)
        mag_clf_c.fit(Xc_tr, is_mag_c)
        proba_c = (mag_clf_c.predict_proba(Xc_te)[:, 1]
                   if hasattr(mag_clf_c, "predict_proba")
                   else mag_clf_c.predict(Xc_te).astype(float))
        mag_flag_c = proba_c >= 0.45
        Xc_tr_mag = np.column_stack([Xc_tr, oc_tr.reshape(-1, 1)])
        Xc_te_mag = np.column_stack([Xc_te, oc_te_pred_cv.reshape(-1, 1)])
        yp_m = np.zeros(len(cv_te_i))
        mag_idx_c = np.where(mag_flag_c)[0]
        if mag_idx_c.size > 0 and is_mag_c.sum() >= 5:
            mr_c = xgb.XGBRegressor(n_estimators=200, learning_rate=0.05, max_depth=4,
                                     random_state=42, n_jobs=-1, verbosity=0)
            mr_c.fit(Xc_tr_mag[is_mag_c.astype(bool)], mc_tr[is_mag_c.astype(bool)])
            yp_m[mag_idx_c] = np.clip(mr_c.predict(Xc_te_mag[mag_idx_c]), 0, None)
        cv_mag_r2.append(r2_score(mc_te, yp_m))
        log.info("  Fold %d: R2(gap)=%.3f  R2(mag)=%.3f",
                  fold_i + 1, cv_gap_r2[-1], cv_mag_r2[-1])

    cv_results = {
        "R2_gap_CV_mean": float(np.mean(cv_gap_r2)),
        "R2_gap_CV_std": float(np.std(cv_gap_r2)),
        "R2_mag_CV_mean": float(np.mean(cv_mag_r2)),
        "R2_mag_CV_std": float(np.std(cv_mag_r2)),
        "n_train": n_tr, "n_test": n_test,
        "R2_gap_test": r2_gap, "MAE_gap_test": mae_gap,
        "R2_mag_test": r2_mag, "MAE_mag_test": mae_mag,
        "ordering_accuracy_test": ord_acc,
    }
    log.info("R2(gap) CV: %.3f +/- %.3f", cv_results["R2_gap_CV_mean"], cv_results["R2_gap_CV_std"])
    log.info("R2(mag) CV: %.3f +/- %.3f", cv_results["R2_mag_CV_mean"], cv_results["R2_mag_CV_std"])

    # ════════════════════════════════════════════════════════════════════
    #  BOOTSTRAP PREDICTION INTERVALS (90%)
    # ════════════════════════════════════════════════════════════════════
    log.info("Computing bootstrap prediction intervals (%d resamples) ...", n_bootstrap)
    alpha = (1 - ci_alpha) / 2
    gap_boot = np.zeros((n_bootstrap, n_test))
    mag_boot = np.zeros((n_bootstrap, n_test))

    for b in range(n_bootstrap):
        rng = np.random.default_rng(b)
        idx = rng.choice(n_tr, size=n_tr, replace=True)
        Xb, gb, mb, ob = X_tr[idx], g_tr[idx], m_tr[idx], o_tr[idx]
        ins_b = gb >= 0.01
        if ins_b.sum() >= 4:
            rf_b = RandomForestRegressor(n_estimators=100, random_state=b, n_jobs=-1)
            rf_b.fit(Xb[ins_b], np.log1p(gb[ins_b]))
            is_ins_b = (clf_metal.predict(X_te).astype(bool)
                        if clf_metal else np.ones(n_test, bool))
            gp_b = np.zeros(n_test)
            if is_ins_b.sum():
                gp_b[is_ins_b] = np.expm1(rf_b.predict(X_te[is_ins_b]))
            gap_boot[b] = np.clip(gp_b, 0, None)
        is_mag_b = (mb > 0.5)
        if is_mag_b.sum() >= 4:
            Xb_mag = np.column_stack([Xb, ob.reshape(-1, 1)])
            rf_mb = RandomForestRegressor(n_estimators=100, random_state=b, n_jobs=-1)
            rf_mb.fit(Xb_mag[is_mag_b], mb[is_mag_b])
            mp_b = np.zeros(n_test)
            proba_b = clf_nm_mag.predict_proba(X_te)[:, 1]
            mag_b_idx = np.where(proba_b >= 0.45)[0]
            if mag_b_idx.size:
                mp_b[mag_b_idx] = np.clip(rf_mb.predict(X_te_mag_pred[mag_b_idx]), 0, None)
            mag_boot[b] = mp_b

        if (b + 1) % 10 == 0:
            log.info("  Bootstrap %d/%d done", b + 1, n_bootstrap)

    gap_lo = np.quantile(gap_boot, alpha, axis=0)
    gap_hi = np.quantile(gap_boot, 1 - alpha, axis=0)
    mag_lo = np.quantile(mag_boot, alpha, axis=0)
    mag_hi = np.quantile(mag_boot, 1 - alpha, axis=0)

    gap_cov = float(np.mean((g_te >= gap_lo) & (g_te <= gap_hi)))
    mag_cov = float(np.mean((m_te >= mag_lo) & (m_te <= mag_hi)))
    log.info("90%% CI coverage — gap: %.2f%%  mag: %.2f%%", gap_cov * 100, mag_cov * 100)
    cv_results["gap_CI_coverage"] = gap_cov
    cv_results["mag_CI_coverage"] = mag_cov

    # ════════════════════════════════════════════════════════════════════
    #  FEATURE IMPORTANCE
    # ════════════════════════════════════════════════════════════════════
    imp_vals = np.zeros(len(FEAT))
    if "all" in gap_models:
        imp_vals += gap_models["all"]["rf"].feature_importances_
    if rf_mag is not None:
        imp_vals += rf_mag.feature_importances_[:len(FEAT)]
    if clf_ord is not None:
        imp_vals += clf_ord.feature_importances_
    imp_vals /= imp_vals.sum()
    imp = pd.DataFrame({"Feature": FEAT, "Importance": imp_vals}) \
        .sort_values("Importance", ascending=False).reset_index(drop=True)

    # ════════════════════════════════════════════════════════════════════
    #  SAVE MODELS
    # ════════════════════════════════════════════════════════════════════
    model_bundle = {
        "clf_metal": clf_metal, "clf_ord": clf_ord, "clf_nm_mag": clf_nm_mag,
        "gap_models": gap_models, "iso_gap": iso_gap,
        "rf_mag": rf_mag, "xgb_mag": xgb_mag, "hgb_mag": hgb_mag,
        "features": FEAT,
    }
    mpath = os.path.join(OUTPUT_DIR, "mxcy_model_bundle_v6.pkl")
    joblib.dump(model_bundle, mpath)
    log.info("Model bundle saved -> %s", mpath)

    # ════════════════════════════════════════════════════════════════════
    #  BUILD TEST RESULT DATAFRAME
    # ════════════════════════════════════════════════════════════════════
    ord_labels = {0: "NM", 1: "AFM", 2: "FM"}
    df_te_out = df_te[["Compound", "M", "C", "Stoichiometry"]].copy()
    df_te_out["Gap_Real_eV"] = g_te
    df_te_out["Gap_Pred_eV"] = gap_te_pred.round(4)
    df_te_out["Gap_Error_eV"] = (gap_te_pred - g_te).round(4)
    df_te_out["Gap_CI_lo"] = gap_lo.round(4)
    df_te_out["Gap_CI_hi"] = gap_hi.round(4)
    df_te_out["Mag_Real_muB"] = m_te
    df_te_out["Mag_Pred_muB"] = mag_te_pred.round(4)
    df_te_out["Mag_Error_muB"] = (mag_te_pred - m_te).round(4)
    df_te_out["Mag_CI_lo"] = mag_lo.round(4)
    df_te_out["Mag_CI_hi"] = mag_hi.round(4)
    df_te_out["Ordering_Real"] = [ord_labels.get(o, "?") for o in o_te]
    df_te_out["Ordering_Pred"] = ([ord_labels.get(o, "?") for o in clf_ord.predict(X_te)]
                                   if clf_ord else ["?" for _ in range(n_test)])
    return imp, cv_results, mpath, df_te_out


log.info("train_ml_v6() defined.")

# =============================================================================
#  SHAP FEATURE IMPORTANCE
# =============================================================================

def run_shap_analysis(df: pd.DataFrame, model_path: str) -> str | None:
    if not (SHAP_OK and model_path):
        log.info("SHAP analysis skipped (shap not installed or model not trained).")
        return None

    bundle = joblib.load(model_path)
    if "gap_models" not in bundle:
        log.warning("Gap model not found in bundle — skipping SHAP.")
        return None

    rf_all = bundle["gap_models"].get("all", {}).get("rf")
    df_real = df[df.Is_Real].reset_index(drop=True)
    X_all = df_real[FEATURES_V6].values

    if rf_all is None or len(X_all) == 0:
        log.warning("No trained gap model / no real data — skipping SHAP.")
        return None

    log.info("Running SHAP TreeExplainer on band gap model ...")
    explainer = shap.TreeExplainer(rf_all)
    shap_vals = explainer.shap_values(X_all)

    plt.figure(figsize=(9, 6))
    shap.summary_plot(shap_vals, X_all, feature_names=FEATURES_V6,
                       show=False, max_display=20)
    plt.title("SHAP summary — band gap model (v6.0)", fontsize=11)
    plt.tight_layout()
    fpath = os.path.join(OUTPUT_DIR, "shap_gap_v6.png")
    plt.savefig(fpath, dpi=150, bbox_inches="tight")
    plt.close()
    log.info("SHAP plot saved -> %s", fpath)
    return fpath


# =============================================================================
#  EXCEL EXPORT
# =============================================================================

def export_excel_v6(df, imp, cv_results, df_test=None) -> str:
    wb = Workbook()

    ws = wb.active; ws.title = "Dataset v6"
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    if not imp.empty:
        ws2 = wb.create_sheet("Feature Importance (v6)")
        for row in dataframe_to_rows(imp, index=False, header=True):
            ws2.append(row)

    ws3 = wb.create_sheet("ML CV Results")
    ws3.append(["Metric", "Value"])
    for k, v in cv_results.items():
        ws3.append([k, round(v, 4) if isinstance(v, float) else v])

    ws4 = wb.create_sheet("Half-Metal Predictions")
    hm_df = df[df["Is_Half_Metal"]].copy()
    for row in dataframe_to_rows(
            hm_df[["Compound", "M", "C", "Stoichiometry",
                   "Spin_Polarization", "U_W_ratio", "exchange_corr_ratio",
                   "zsa_delta", "stoner_split"]].reset_index(drop=True),
            index=False, header=True):
        ws4.append(row)

    if df_test is not None and not df_test.empty:
        ws5 = wb.create_sheet("Prediction_vs_Real")
        ws5.append(["Prediction vs Real with 90% CI - Held-Out Test Set (v6.0)"])
        ws5.append([])
        for row in dataframe_to_rows(df_test, index=False, header=True):
            ws5.append(row)

    rationale = [
        ["Feature", "Physics Basis", "Expected Effect", "Source", "Version"],
        ["d_electrons", "Crystal field / Hund's rule", "More filled -> smaller gap", "Thesis Ch.I", "v4"],
        ["Hubbard_U", "Mott criterion (on-site U)", "Larger U -> Mott gap", "Article Eq.1-2", "v4"],
        ["chi_diff", "Electronegativity -> ionicity", "Larger dChi -> larger gap", "Thesis Tab.II.3", "v4"],
        ["bond_length_A", "Orbital overlap -> bandwidth", "Longer d -> narrower W", "Thesis Tab.II.4", "v4"],
        ["bandwidth_W", "d-band width / itinerancy", "Wider W -> metallic", "Article Sec.4.2", "v4"],
        ["U_W_ratio", "Mott criterion U/W", "U/W > 1 -> insulating", "Thesis Ch.II.3", "v4"],
        ["bond_angle_deg", "Goodenough-Kanamori", "180deg -> AFM, 90deg -> FM", "Thesis Ch.II.4", "v4"],
        ["ca_ratio", "c/a -> interlayer distance", "c/a up -> gap down", "Article Fig.9", "v4"],
        ["exchange_corr_ratio", "Stoner I/W", "I/W > 1 -> FM", "Article Sec.4.2", "v4"],
        ["soc_lambda", "Spin-orbit coupling (meV)", "Large lambda -> quenches magnetism", "Abragam 1970", "v5"],
        ["crystal_field_10Dq", "Crystal-field splitting", "Large 10Dq -> t2g/eg gap", "Orgel", "v5"],
        ["ionicity", "Phillips ionicity", "More ionic -> larger gap", "Phillips 1970", "v5"],
        ["madelung_energy", "Born-Madelung ionic energy", "Larger E_Mad -> ionic -> gap up", "Born-Madelung", "v5"],
        ["valence_electron_count", "VEC = x*nd + y*6", "Certain VEC -> insulating", "Hume-Rothery", "v5"],
        ["nd_half_fill", "Hund half-fill proximity", "Half-fill -> max Hund moment", "Hund rules", "v5"],
        ["r_ionic_ratio", "Ionic radius ratio rM/rC", "Mismatch -> distortion -> gap", "Goldschmidt", "v5"],
        ["anion_polarizability", "Chalcogen polarizability", "Larger alpha -> covalent -> gap down", "Handbook", "v5"],
        ["jahn_teller_active", "JT-active d configs", "JT -> distortion -> gap opens", "JT theorem", "v5"],
        ["superexchange_factor", "sin^2(t/2)cos^2(t/2) GK", "Peak at 90deg (FM), 0 at 180deg", "Goodenough 1955", "v5"],
        ["is_oxide", "Binary anion type flag", "Oxides: large gaps", "-", "v5"],
        ["is_3d_metal", "Binary period flag", "3d: stronger exchange", "-", "v5"],
        ["is_5d_metal", "Binary period flag", "5d: strong SOC, quench mag", "-", "v5"],
        ["zsa_delta", "Zaanen-Sawatzky-Allen Delta", "Delta < 0 -> charge-transfer ins.", "ZSA 1985", "v6 NEW"],
        ["soc_sq", "lambda^2 second-order SOC", "Quenches orbital moment", "Abragam 1970", "v6 NEW"],
        ["stoner_split", "Stoner I x W exchange splitting", "E_ex > 1 -> FM instability", "Stoner 1938", "v6 NEW"],
        ["d_fill_ratio", "n_d / 10 normalised d-filling", "0.5 = half-fill max moment", "Hund rules", "v6 NEW"],
        ["goldschmidt_t", "Ionic radius tolerance factor", "t != 1 -> distortion -> gap mod.", "Goldschmidt 1926", "v6 NEW"],
    ]
    ws6 = wb.create_sheet("Feature Rationale v6")
    for r in rationale:
        ws6.append(r)

    path = os.path.join(OUTPUT_DIR, "MxCy_v6_Full_Results.xlsx")
    wb.save(path)
    log.info("Excel saved -> %s", path)
    return path


# =============================================================================
#  PLOTS
# =============================================================================

def make_plots_v6(df, imp, cv_results, df_test) -> str:
    """Generate comparison plots: gap, mag, ordering, version comparison, CI."""
    fig, axes = plt.subplots(2, 3, figsize=(15, 9))
    fig.suptitle("MxCy v6.0 - ML Results", fontsize=13, fontweight="bold")

    ax = axes[0, 0]
    if df_test is not None and not df_test.empty:
        ax.scatter(df_test["Gap_Real_eV"], df_test["Gap_Pred_eV"],
                   alpha=0.6, color="#378ADD", s=30)
        lim = max(df_test["Gap_Real_eV"].max(), df_test["Gap_Pred_eV"].max()) * 1.05
        ax.plot([0, lim], [0, lim], "k--", lw=0.8, alpha=0.5)
        r2 = cv_results.get("R2_gap_test", float("nan"))
        mae = cv_results.get("MAE_gap_test", float("nan"))
        ax.set_title(f"Band gap: pred vs real\nR2={r2:.3f}  MAE={mae:.3f} eV", fontsize=10)
        ax.set_xlabel("Real (eV)"); ax.set_ylabel("Predicted (eV)")

    ax = axes[0, 1]
    if df_test is not None and not df_test.empty:
        ax.scatter(df_test["Mag_Real_muB"], df_test["Mag_Pred_muB"],
                   alpha=0.6, color="#1D9E75", s=30)
        lim = max(df_test["Mag_Real_muB"].max(), df_test["Mag_Pred_muB"].max()) * 1.05 + 1
        ax.plot([0, lim], [0, lim], "k--", lw=0.8, alpha=0.5)
        r2 = cv_results.get("R2_mag_test", float("nan"))
        mae = cv_results.get("MAE_mag_test", float("nan"))
        ax.set_title(f"Magnetization: pred vs real\nR2={r2:.3f}  MAE={mae:.3f} muB", fontsize=10)
        ax.set_xlabel("Real (muB)"); ax.set_ylabel("Predicted (muB)")

    ax = axes[0, 2]
    top = imp.head(15)
    ax.barh(top["Feature"][::-1], top["Importance"][::-1], color="#7F77DD")
    ax.set_title("Feature importance (top 15)", fontsize=10)
    ax.set_xlabel("Mean importance")
    ax.tick_params(axis="y", labelsize=8)

    ax = axes[1, 0]
    versions = ["v4.0", "v5.1", "v6.0"]
    r2_gap_vals = [0.517, 0.563, cv_results.get("R2_gap_test", 0)]
    r2_mag_vals = [0.718, -0.754, cv_results.get("R2_mag_test", 0)]
    x = np.arange(len(versions))
    w = 0.35
    ax.bar(x - w / 2, r2_gap_vals, w, label="R2(gap)", color="#378ADD")
    ax.bar(x + w / 2, r2_mag_vals, w, label="R2(mag)", color="#1D9E75")
    ax.axhline(0, color="k", lw=0.7, ls="--")
    ax.set_xticks(x); ax.set_xticklabels(versions)
    ax.set_title("R2 (test): version comparison", fontsize=10)
    ax.legend(fontsize=8)

    ax = axes[1, 1]
    if df_test is not None and "Gap_CI_lo" in df_test.columns:
        ci_width = (df_test["Gap_CI_hi"] - df_test["Gap_CI_lo"]).dropna()
        ax.hist(ci_width, bins=15, color="#B5D4F4", edgecolor="white")
        ax.axvline(ci_width.median(), color="#185FA5", lw=1.5, ls="--",
                   label=f"median={ci_width.median():.2f} eV")
        cov = cv_results.get("gap_CI_coverage", float("nan"))
        ax.set_title(f"90% CI width - band gap\ncoverage={cov:.1%}", fontsize=10)
        ax.set_xlabel("CI width (eV)"); ax.set_ylabel("Count")
        ax.legend(fontsize=8)

    ax = axes[1, 2]
    if df_test is not None and "Ordering_Real" in df_test.columns:
        labels = ["NM", "AFM", "FM"]
        real_counts = [df_test["Ordering_Real"].eq(l).sum() for l in labels]
        pred_counts = [df_test["Ordering_Pred"].eq(l).sum() for l in labels]
        x2 = np.arange(len(labels))
        ax.bar(x2 - w / 2, real_counts, w, label="Real", color="#B5D4F4")
        ax.bar(x2 + w / 2, pred_counts, w, label="Pred", color="#9FE1CB")
        ax.set_xticks(x2); ax.set_xticklabels(labels)
        acc = cv_results.get("ordering_accuracy_test", float("nan"))
        ax.set_title(f"Ordering distribution\ntest acc={acc:.2%}", fontsize=10)
        ax.legend(fontsize=8)

    plt.tight_layout()
    fpath = os.path.join(OUTPUT_DIR, "MxCy_v6_summary.png")
    plt.savefig(fpath, dpi=150, bbox_inches="tight")
    plt.close()
    log.info("Summary plot saved -> %s", fpath)
    return fpath

# =============================================================================
#  MAIN ENTRY POINT
# =============================================================================

def main():
    t_pipeline_start = time.time()

    log.info("")
    log.info("STEP 1/7 — Configuration check")
    log.info("  OUTPUT_DIR    = %s", OUTPUT_DIR)
    log.info("  MP_API_KEY    = %s", "set (%d chars)" % len(MP_API_KEY) if MP_API_KEY else "NOT SET")
    log.info("  PAPER_DIR     = %s", PAPER_DIR or "(not set — PDF mining disabled)")
    log.info("  PAPER_DB_CSV  = %s", PAPER_DB_CSV_PATH or "(not set)")

    if not MP_API_KEY:
        log.warning(
            "MP_API_KEY is not set. The script will still run and build an "
            "empirical-only dataset, but accuracy will be much lower. "
            "Set the MP_API_KEY environment variable to fetch real data."
        )

    # ── STEP 2: PDF literature mining (optional, only if PAPER_DIR is set) ──
    log.info("")
    log.info("STEP 2/7 — PDF literature mining")
    pdf_records = run_pdf_mining_stage()
    # Note: pdf_records are written to pdf_mining_candidates_v6.csv for
    # manual review. They are NOT auto-imported into PAPER_DB_MANUAL —
    # this is a deliberate safety gate against false positives at scale.
    # To use them, review the CSV, then set PAPER_DB_CSV_PATH to a filtered
    # copy (or paste confirmed rows into PAPER_DB_MANUAL above) and re-run.
    if pdf_records:
        log.info(
            "%d candidate rows extracted from PDFs. Review "
            "pdf_mining_candidates_v6.csv before the next run to "
            "incorporate them into the paper database.",
            len(pdf_records),
        )

    # ── STEP 3: Load paper database (manual + reviewed CSV) ─────────────────
    log.info("")
    log.info("STEP 3/7 — Loading paper database")
    paper_db = load_paper_db()

    # ── STEP 4: Fetch Materials Project data ─────────────────────────────────
    log.info("")
    log.info("STEP 4/7 — Fetching Materials Project data")
    real_dict = fetch_real_data_v6(MP_API_KEY, paper_db=paper_db)

    # ── STEP 5: Build full dataset ───────────────────────────────────────────
    log.info("")
    log.info("STEP 5/7 — Building full feature dataset")
    df = build_full_dataset_v6(real_dict)
    dataset_csv = os.path.join(OUTPUT_DIR, "MxCy_dataset_v6_raw.csv")
    df.to_csv(dataset_csv, index=False)
    log.info("Raw dataset saved -> %s", dataset_csv)

    # ── STEP 6: Train ML models ──────────────────────────────────────────────
    log.info("")
    log.info("STEP 6/7 — Training ML models")
    feat_imp, cv_results, model_path, df_test = train_ml_v6(df)

    if model_path is None:
        log.warning(
            "Training was skipped (insufficient real data). "
            "Increase MAX_HULL, add more paper data, or check your MP_API_KEY."
        )
    else:
        # SHAP
        log.info("")
        log.info("STEP 7/7 — SHAP, Excel export, and plots")
        run_shap_analysis(df, model_path)

        # Excel export
        excel_path = export_excel_v6(df, feat_imp, cv_results, df_test)

        # Plots
        plot_path = make_plots_v6(df, feat_imp, cv_results, df_test)

        # Zip everything for easy retrieval
        import zipfile
        zip_path = os.path.join(OUTPUT_DIR, "MxCy_results_v6.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in os.listdir(OUTPUT_DIR):
                if f.endswith((".pkl", ".xlsx", ".png", ".csv", ".json")):
                    zf.write(os.path.join(OUTPUT_DIR, f), f)
        log.info("Results ZIP -> %s", zip_path)

    elapsed = time.time() - t_pipeline_start
    log.info("")
    log.info("=" * 70)
    log.info("Pipeline run complete in %.1f minutes.", elapsed / 60.0)
    log.info("All outputs in: %s", os.path.abspath(OUTPUT_DIR))
    log.info("=" * 70)

    return {
        "dataset": df,
        "feature_importance": feat_imp,
        "cv_results": cv_results,
        "model_path": model_path,
        "test_predictions": df_test,
    }


if __name__ == "__main__":
    results = main()
