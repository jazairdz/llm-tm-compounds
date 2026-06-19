"""
PDF Library Organizer  —  with Gemini AI classification
========================================================
Scans a local folder of PDF papers, extracts the real title from inside
each file, renames it, classifies it using Gemini AI into one of 4
categories, moves it to the matching subfolder, and exports an Excel report.

Usage
-----
    python pdf_organizer.py "G:\\MyResearch\\AllPapers"

You will be asked for your free Gemini API key on first run.
Get one at: https://aistudio.google.com  (no credit card needed)

Categories created automatically inside PAPER_DIR:
    0_Unclassified
    1_Out_of_Subject
    2_Physics_Books
    3_Manuscripts
    4_Subject_Papers

Requirements
------------
    pip install pypdf openpyxl google-generativeai
"""

import argparse
import re
import shutil
import subprocess
import sys
import time
import unicodedata
from pathlib import Path


# ---------------------------------------------------------------------------
# 0. Dependency check
# ---------------------------------------------------------------------------

def _install(pkg):
    print(f"  Installing {pkg} ...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

def ensure_dependencies():
    for pkg, imp in [("pypdf","pypdf"), ("openpyxl","openpyxl"), ("google-generativeai","google.generativeai")]:
        try:
            __import__(imp)
        except ImportError:
            _install(pkg)


# ---------------------------------------------------------------------------
# 1. Gemini AI setup
# ---------------------------------------------------------------------------

GEMINI_MODEL   = "gemini-1.5-flash"
API_KEY_FILE   = Path.home() / ".pdf_organizer_gemini_key"
RATE_LIMIT_SEC = 4.5        # ~13 RPM — free tier allows 15 RPM

_gemini_client = None

def _load_api_key():
    if API_KEY_FILE.exists():
        key = API_KEY_FILE.read_text().strip()
        if key:
            return key
    print("\n" + "=" * 60)
    print("  Gemini API key not found.")
    print("  Get your FREE key at: https://aistudio.google.com")
    print("  (Sign in with Google -> Create API key -> Copy it)")
    print("=" * 60)
    key = input("\n  Paste your Gemini API key here: ").strip()
    if not key:
        print("[ERROR] No API key provided. Exiting.")
        sys.exit(1)
    API_KEY_FILE.write_text(key)
    print(f"  Key saved to {API_KEY_FILE}  (won't be asked again)\n")
    return key

def get_gemini_client():
    global _gemini_client
    if _gemini_client is None:
        import google.generativeai as genai
        genai.configure(api_key=_load_api_key())
        _gemini_client = genai.GenerativeModel(GEMINI_MODEL)
    return _gemini_client


# ---------------------------------------------------------------------------
# 2. Title extraction
# ---------------------------------------------------------------------------

def _sanitize_filename(title, max_len=120):
    title = unicodedata.normalize("NFKD", title)
    title = title.encode("ascii", "ignore").decode("ascii")
    title = re.sub(r'[\\/:*?"<>|]', "", title)
    title = re.sub(r"\s+", " ", title).strip()
    return title[:max_len] if title else ""

def extract_title(pdf_path):
    import pypdf
    try:
        reader = pypdf.PdfReader(str(pdf_path))
        meta = reader.metadata
        if meta and getattr(meta, "title", None):
            t = meta.title.strip()
            if len(t) > 5 and not re.fullmatch(r"[\w\-]+", t):
                return t
        if reader.pages:
            text = reader.pages[0].extract_text() or ""
            lines = [l.strip() for l in text.splitlines() if l.strip()]
            for line in lines[:15]:
                if 8 <= len(line) <= 150 and re.search(r"[a-zA-Z]{3}", line):
                    skip = re.search(
                        r"^(abstract|introduction|doi|http|www|journal|vol\.?\s*\d|"
                        r"received|accepted|published|copyright|issn|isbn|page|\d+$)",
                        line, re.IGNORECASE)
                    if not skip:
                        return line
    except Exception:
        pass
    return pdf_path.stem

def extract_abstract(pdf_path, max_chars=800):
    import pypdf
    try:
        reader = pypdf.PdfReader(str(pdf_path))
        text = ""
        for page in reader.pages[:3]:
            text += page.extract_text() or ""
        m = re.search(r"abstract[\s\S]{0,20}?\n([\s\S]{50,800})", text, re.IGNORECASE)
        if m:
            return m.group(1).replace("\n", " ").strip()
        if reader.pages:
            return (reader.pages[0].extract_text() or "")[:max_chars]
    except Exception:
        pass
    return ""

def rename_pdf(pdf_path, title):
    safe = _sanitize_filename(title)
    if not safe:
        return pdf_path
    new_path = pdf_path.parent / (safe + ".pdf")
    if new_path == pdf_path:
        return pdf_path
    counter = 1
    while new_path.exists() and new_path != pdf_path:
        new_path = pdf_path.parent / (safe + f"_{counter}.pdf")
        counter += 1
    pdf_path.rename(new_path)
    return new_path


# ---------------------------------------------------------------------------
# 3. Classification
# ---------------------------------------------------------------------------

CLASSIFICATION_PROMPT = """You are a research librarian organizing a physics / materials science library.

Classify the document below into EXACTLY ONE of these 4 categories:

  subject_paper  — Research paper on transition metal chalcogenides (MxCy compounds:
                   oxides, sulfides, selenides, tellurides), electronic structure,
                   band gap, magnetic properties, DFT/GGA/LDA calculations, etc.

  physics_book   — Textbook or reference book on general physics topics
                   (solid state physics, quantum mechanics, thermodynamics, etc.)

  manuscript     — PhD thesis, dissertation, or master memoir.

  out_of_subject — Anything unrelated: biology, chemistry, machine learning, etc.

Respond with ONLY the category key. Nothing else.
Valid responses: subject_paper | physics_book | manuscript | out_of_subject

Title: {title}
Abstract: {abstract}
"""

VALID_CATS = {"subject_paper", "physics_book", "manuscript", "out_of_subject"}

# Keyword fallback used when API fails
_SUBJECT_KW    = ["chalcogenide","sulfide","selenide","telluride","oxide","band gap",
                  "magnetization","DFT","GGA","LDA","VASP","Wien2k","ab initio",
                  "first-principles","transition metal","half-metal","spin polariz",
                  "Hubbard","Mott","ferromagnet","antiferromagnet"]
_BOOK_KW       = ["textbook","introduction to","fundamentals","principles of",
                  "solid state physics","quantum mechanics","second edition",
                  "third edition","isbn","table of contents"]
_MANUSCRIPT_KW = ["thesis","dissertation","doctorat","phd","master","magistere",
                  "magister","memoire","submitted to","degree of doctor"]
_OUT_KW        = ["biology","polymer","organic","machine learning","deep learning",
                  "economics","finance","climate","medicine","pharmacol"]

def _keyword_classify(title, abstract):
    t = (title + " " + abstract).lower()
    scores = {
        "subject_paper":  sum(1 for kw in _SUBJECT_KW    if kw.lower() in t),
        "physics_book":   sum(1 for kw in _BOOK_KW       if kw.lower() in t),
        "manuscript":     sum(1 for kw in _MANUSCRIPT_KW if kw.lower() in t),
        "out_of_subject": sum(1 for kw in _OUT_KW        if kw.lower() in t),
    }
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "unclassified"

def classify(title, abstract):
    """Classify using Gemini AI, fall back to keywords on error."""
    try:
        client = get_gemini_client()
        prompt = CLASSIFICATION_PROMPT.format(
            title=title[:300], abstract=abstract[:600])
        response = client.generate_content(prompt)
        cat = response.text.strip().lower().replace(" ", "_")
        if cat in VALID_CATS:
            return cat, "gemini"
        for valid in VALID_CATS:
            if valid in cat:
                return valid, "gemini"
        return _keyword_classify(title, abstract), "keyword_fallback"
    except Exception as e:
        err = str(e)
        if "429" in err or "quota" in err.lower():
            print("\n  [!] Rate limit hit — waiting 30s ...")
            time.sleep(30)
            return classify(title, abstract)   # retry
        print(f"  [!] Gemini error: {e} — using keyword fallback")
        return _keyword_classify(title, abstract), "keyword_fallback"


# ---------------------------------------------------------------------------
# 4. Excel report
# ---------------------------------------------------------------------------

CAT_FOLDERS = {
    "subject_paper":  "4_Subject_Papers",
    "physics_book":   "2_Physics_Books",
    "manuscript":     "3_Manuscripts",
    "out_of_subject": "1_Out_of_Subject",
    "unclassified":   "0_Unclassified",
}

CAT_COLORS = {
    "subject_paper":  "D9EAD3",
    "physics_book":   "FFF2CC",
    "manuscript":     "FCE5CD",
    "out_of_subject": "F4CCCC",
    "unclassified":   "EFEFEF",
}

def write_excel_report(records, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "All Papers"

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    headers = ["#","Original Filename","New Title","Category",
               "Method","Folder","Abstract Snippet","Final Path"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, rec in enumerate(records, 1):
        cat    = rec.get("category", "unclassified")
        folder = CAT_FOLDERS.get(cat, CAT_FOLDERS["unclassified"])
        ws.append([
            i,
            rec["original_filename"],
            rec["title"],
            cat,
            rec.get("method", ""),
            folder,
            rec["abstract_snippet"][:200],
            rec.get("final_path", ""),
        ])
        fill = PatternFill("solid", fgColor=CAT_COLORS.get(cat, "FFFFFF"))
        for col in range(1, len(headers) + 1):
            ws.cell(i + 1, col).fill = fill

    for col, w in enumerate([5,45,70,20,18,30,60,60], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Category","Folder","Total","Gemini AI","Keyword fallback"])
    for cat, folder in CAT_FOLDERS.items():
        total  = sum(1 for r in records if r.get("category") == cat)
        ai_cnt = sum(1 for r in records if r.get("category") == cat and r.get("method") == "gemini")
        ws2.append([cat, folder, total, ai_cnt, total - ai_cnt])

    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# 5. Main
# ---------------------------------------------------------------------------

def main(paper_dir):
    paper_path = Path(paper_dir)

    if not paper_path.exists():
        print(f"[ERROR] Directory not found: {paper_path}")
        sys.exit(1)
    if not paper_path.is_dir():
        print(f"[ERROR] Not a directory: {paper_path}")
        sys.exit(1)

    pdf_files = sorted(paper_path.glob("*.pdf"))
    if not pdf_files:
        print(f"[WARNING] No PDF files found in: {paper_path}")
        sys.exit(0)

    print(f"\n{'='*60}")
    print(f"  PDF Library Organizer — Gemini AI")
    print(f"  Folder : {paper_path}")
    print(f"  Files  : {len(pdf_files)} PDF(s)")
    print(f"{'='*60}\n")

    # Trigger API key setup before processing starts
    get_gemini_client()

    # ── Step 1: Extract titles & rename ────────────────────────────────
    print(f"[1/3] Extracting titles and renaming files ...\n")
    records = []
    for i, pdf in enumerate(pdf_files, 1):
        title    = extract_title(pdf)
        abstract = extract_abstract(pdf)
        new_pdf  = rename_pdf(pdf, title)
        records.append({
            "original_filename": pdf.name,
            "title":             title,
            "new_filename":      new_pdf.name,
            "abstract_snippet":  abstract[:200],
            "path":              new_pdf,
            "category":          None,
            "method":            None,
            "final_path":        "",
        })
        print(f"  [{i:3d}/{len(pdf_files)}] {new_pdf.name[:70]}")

    # ── Step 2: Classify ───────────────────────────────────────────────
    print(f"\n[2/3] Classifying with Gemini AI ...\n")
    for i, rec in enumerate(records, 1):
        cat, method = classify(rec["title"], rec["abstract_snippet"])
        rec["category"] = cat
        rec["method"]   = method
        tag = "[AI]" if method == "gemini" else "[KW]"
        print(f"  [{i:3d}/{len(records)}] {tag} {rec['title'][:55]:55s}  ->  {cat}")
        if method == "gemini" and i < len(records):
            time.sleep(RATE_LIMIT_SEC)

    print(f"\n  Summary:")
    for cat, folder in CAT_FOLDERS.items():
        count = sum(1 for r in records if r.get("category") == cat)
        if count:
            print(f"    {folder:35s}  {count:4d} file(s)")

    # ── Step 3: Move files ─────────────────────────────────────────────
    print(f"\n[3/3] Moving files ...\n")
    for folder in CAT_FOLDERS.values():
        (paper_path / folder).mkdir(exist_ok=True)

    errors = []
    for rec in records:
        src    = paper_path / rec["new_filename"]
        folder = CAT_FOLDERS.get(rec["category"], CAT_FOLDERS["unclassified"])
        dst    = paper_path / folder / rec["new_filename"]
        counter = 1
        while dst.exists():
            stem = Path(rec["new_filename"]).stem
            dst  = paper_path / folder / f"{stem}_{counter}.pdf"
            counter += 1
        try:
            if src.exists():
                shutil.move(str(src), str(dst))
                rec["final_path"] = str(dst)
                print(f"  OK  {rec['new_filename'][:60]:60s}  ->  {folder}")
            else:
                rec["final_path"] = "FILE NOT FOUND"
                errors.append(rec["new_filename"])
        except Exception as e:
            rec["final_path"] = f"ERROR: {e}"
            errors.append(rec["new_filename"])
            print(f"  !!  {rec['new_filename']} — {e}")

    # ── Excel report ───────────────────────────────────────────────────
    report_path = paper_path / "PDF_Library_Report.xlsx"
    write_excel_report(records, report_path)

    ai_count = sum(1 for r in records if r.get("method") == "gemini")
    print(f"\n{'='*60}")
    print(f"  Done.  {len(records)} files,  {len(errors)} error(s)")
    print(f"  Gemini AI: {ai_count}   Keyword fallback: {len(records)-ai_count}")
    print(f"  Report -> {report_path}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Organize PDF papers using Gemini AI.")
    parser.add_argument("paper_dir", nargs="?", default="",
        help='Path to your papers folder. e.g. "G:\\MyResearch\\AllPapers"')
    args = parser.parse_args()

    if not args.paper_dir:
        print("PDF Library Organizer — powered by Gemini AI (free)")
        print("=" * 50)
        args.paper_dir = input("Enter path to your papers folder: ").strip().strip('"')

    ensure_dependencies()
    main(args.paper_dir)
